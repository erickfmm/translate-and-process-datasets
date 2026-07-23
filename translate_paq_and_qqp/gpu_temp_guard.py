"""GPU temperature guard - supervisor approach.

A CPU-side supervisor watches the GPU temperature by shelling out to
``nvidia-smi``. It owns the lifecycle of a GPU translation subprocess:

* when the temperature reaches ``temp_max`` it **kills** the GPU subprocess
  (freeing all VRAM - "sin nada") so the hardware can cool down;
* when the temperature drops back to ``temp_resume`` it **restarts** the
  subprocess, resuming from the last row already saved to the output file;
* if the temperature ever reaches ``temp_stop`` it kills the subprocess and
  aborts permanently (safety net).

The supervisor itself runs purely on the CPU and holds no GPU memory, so it
can keep watching while the GPU is idle.

Progress / resume point is read from the output XLSX's ``index`` column, so
the worker just needs to keep flushing the file continuously (which it
already does) and to append to an existing file instead of overwriting it.
"""

from __future__ import annotations

import os
import signal
import subprocess
import sys
import time
from pathlib import Path
from typing import Callable, List, Optional, Tuple


# ---------------------------------------------------------------------------
# nvidia-smi parsing
# ---------------------------------------------------------------------------
def read_gpu_temperature(gpu_index: int = 0) -> Optional[int]:
	"""Return the temperature (Celsius) of GPU *gpu_index*, or ``None``.

	Uses the query form of ``nvidia-smi`` which prints one temperature per
	line (just the integer, no units). Output example for two GPUs::

		72
		68
	"""
	try:
		result = subprocess.run(
			[
				"nvidia-smi",
				"--query-gpu=temperature.gpu",
				"--format=csv,noheader,nounits",
			],
			capture_output=True,
			text=True,
			timeout=10,
		)
	except (FileNotFoundError, subprocess.TimeoutExpired):
		return None

	if result.returncode != 0:
		return None

	lines = [ln.strip() for ln in result.stdout.splitlines() if ln.strip()]
	if gpu_index < 0 or gpu_index >= len(lines):
		return None
	try:
		return int(lines[gpu_index])
	except ValueError:
		return None


# ---------------------------------------------------------------------------
# Resume helpers (operate on the output XLSX)
# ---------------------------------------------------------------------------
def read_last_index(xlsx_path) -> Optional[int]:
	"""Return the highest ``index`` value stored in *xlsx_path*, or ``None``.

	Uses openpyxl read-only mode so it stays cheap even on very large files
	(it only scans the first column). Returns ``None`` if the file is missing
	or contains no data rows.
	"""
	path = Path(xlsx_path)
	if not path.exists():
		return None
	try:
		from openpyxl import load_workbook
	except ImportError:
		return None

	last: Optional[int] = None
	try:
		wb = load_workbook(path, read_only=True)
		ws = wb.active
		# Column A holds the index; row 1 is the header.
		for (value,) in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
			if isinstance(value, (int, float)):
				last = int(value)
		wb.close()
	except Exception:
		return None
	return last


def load_or_create_workbook(xlsx_path, headers: List[str], resume_append: bool = False):
	"""Return ``(workbook, sheet)`` for writing.

	If *resume_append* is true and *xlsx_path* already exists, the existing
	workbook is loaded so new rows are appended (preserving what was already
	saved across supervisor restarts). Otherwise a fresh workbook is created
	with *headers* as the first row.
	"""
	from openpyxl import Workbook

	path = Path(xlsx_path)
	if resume_append and path.exists():
		from openpyxl import load_workbook
		wb = load_workbook(path)
		ws = wb.active
		return wb, ws
	wb = Workbook()
	ws = wb.active
	ws.title = 'Hoja1'
	ws.append(headers)
	return wb, ws


# ---------------------------------------------------------------------------
# Supervisor
# ---------------------------------------------------------------------------
def _kill_process_group(proc: "subprocess.Popen") -> None:
	"""SIGKILL the whole process group of *proc* (the worker + its children)."""
	if proc.poll() is not None:
		return
	try:
		os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
	except (ProcessLookupError, PermissionError):
		try:
			proc.kill()
		except ProcessLookupError:
			pass
	try:
		proc.wait(timeout=10)
	except subprocess.TimeoutExpired:
		pass


def run_temp_guard_supervisor(
	*,
	script_path: str,
	argv: List[str],
	output_xlsx: str,
	initial_skip_rows: int,
	temp_max: int,
	temp_resume: int,
	temp_stop: int,
	check_interval: int,
	gpu_index: int = 0,
	log_fn: Callable[[str], None] = print,
) -> int:
	"""Run the CPU supervisor loop. Returns a process exit code.

	*script_path* + *argv* is re-executed (with ``TRANSLATE_SUPERVISOR_CHILD=1``
	and ``TRANSLATE_SKIP_ROWS=<n>`` in the environment) as the GPU worker. The
	supervisor kills it at ``temp_max``, restarts it at ``temp_resume`` (resuming
	from the last index found in *output_xlsx*), and aborts at ``temp_stop``.
	"""
	env = os.environ.copy()
	env["TRANSLATE_SUPERVISOR_CHILD"] = "1"

	def _temp() -> Optional[int]:
		return read_gpu_temperature(gpu_index)

	def _wait_below(temp_threshold: int, *, on_stop: bool) -> bool:
		"""Block until temp <= temp_threshold. Return False if temp_stop hit."""
		while True:
			t = _temp()
			if t is not None:
				if t >= temp_stop:
					log_fn(
						f"[Supervisor] STOP: {t}C >= {temp_stop}C "
						f"(GPU {gpu_index}) - aborting"
					)
					return False
				if t <= temp_threshold:
					return True
				log_fn(
					f"[Supervisor] cooling: {t}C (target <= {temp_threshold}C)"
				)
			time.sleep(check_interval)

	while True:
		# Don't spawn into an already-hot GPU: cool first.
		t = _temp()
		if t is not None and t >= temp_max:
			log_fn(
				f"[Supervisor] GPU already at {t}C >= {temp_max}C before start; "
				f"cooling to {temp_resume}C"
			)
			if not _wait_below(temp_resume, on_stop=True):
				return 2

		# Compute the resume point from whatever is already saved.
		last = read_last_index(output_xlsx)
		skip = (last + 1) if last is not None else initial_skip_rows
		env["TRANSLATE_SKIP_ROWS"] = str(skip)
		cur = _temp()
		log_fn(
			f"[Supervisor] launching GPU worker from index {skip} "
			f"(temp {cur if cur is not None else '?'}C)"
		)

		proc = subprocess.Popen(
			[sys.executable, "-u", script_path, *argv],
			env=env,
			start_new_session=True,
		)
		killed_by_us = False
		completed = False

		# Monitor loop: watch temperature + worker status.
		while True:
			t = _temp()
			rc = proc.poll()

			if t is not None and t >= temp_stop:
				log_fn(
					f"[Supervisor] STOP: {t}C >= {temp_stop}C - killing worker "
					f"and aborting"
				)
				_kill_process_group(proc)
				return 2

			if t is not None and t >= temp_max:
				log_fn(
					f"[Supervisor] PAUSE: {t}C >= {temp_max}C - killing GPU "
					f"worker to free VRAM"
				)
				_kill_process_group(proc)
				killed_by_us = True
				break

			if rc is not None:
				if killed_by_us:
					break
				if rc == 0:
					completed = True
				else:
					log_fn(
						f"[Supervisor] worker exited unexpectedly (code {rc}); "
						f"aborting"
					)
					return rc
				break

			time.sleep(check_interval)

		if completed:
			log_fn("[Supervisor] worker completed all rows")
			return 0

		# Worker was killed at temp_max: wait for the GPU to cool, then loop.
		if not _wait_below(temp_resume, on_stop=True):
			return 2
		log_fn(f"[Supervisor] RESUME: restarting worker (temp <= {temp_resume}C)")


# ---------------------------------------------------------------------------
# Smoke test: python gpu_temp_guard.py
# ---------------------------------------------------------------------------
if __name__ == "__main__":
	def _now() -> str:
		return time.strftime("%H:%M:%S")

	print(f"{_now()} temp = {read_gpu_temperature(0)}C")
